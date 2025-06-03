import pandas as pd
import openpyxl
import glob
import os
from openpyxl.styles import PatternFill
from unicodedata import normalize
os.environ["TF_CPP_MIN_LOG_LEVEL"]="0"
# df = pd.read_excel(file_path, engine='openpyxl',usecols=list(range(0,7)),sheet_name="Sheet1",dtype=column_dtype)
neon_green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
def normalize_and_replace(text):
    text = normalize("NFKC", text)
    # Manually replace "ی" with "ي" and "ک" with "ك"
    text = text.replace('ی', 'ي').replace('ک', 'ك')
    # text=text.strip()
    return text
def find_state_name_in_sheet(path,defective_state_name):
    state_name_in_sheet=""
# Create an ExcelFile object first
    excel_file = pd.ExcelFile(path)
    
    # Get all sheet names
    sheet_names = excel_file.sheet_names
    # sheet_names_normalized=[normalize_and_replace(i) for i in sheet_names]
    for i in sheet_names:
        if defective_state_name in normalize_and_replace(i):
            state_name_in_sheet=i
            break
    return (state_name_in_sheet)
def to_number(value):
    
    try:
        return int(value)
    except (ValueError,TypeError):
        try:
            return float(value)
        except(ValueError,TypeError):
            return str(value)
def condition_check(df1,state_name):
    name_exist_in_sheet_for_compare=normalize_and_replace(str(state_name).strip())

    df1_normalized=df1.iloc[:,0].apply(lambda x:normalize_and_replace(str(x).strip()))
    condition = df1_normalized == name_exist_in_sheet_for_compare
    
        
    if condition.any():
        row_index = condition.idxmax()
    else:
        row_index = None
    return (row_index)
current_dir=os.path.dirname(os.path.abspath(__file__))
sub_dir=current_dir+"\\مستندات"
kk=0
df1 = pd.read_excel(current_dir+"\\"+"زیرساخت عملکرد یکساله استانها1403.xlsx",usecols=[1],sheet_name=0,dtype=str)
output_path=current_dir+"\\"+"output.xlsx"
wb = openpyxl.load_workbook(current_dir+"\\"+"زیرساخت عملکرد یکساله استانها1403.xlsx")
ws = wb.worksheets[0]
for i in glob.glob(sub_dir+"\\*"):
    sho=0
    state_name=(i.split("\\"))[-1].strip()
    state_path_xls=sub_dir+"\\"+state_name+"\\*.xls"
    state_path_xlsx=sub_dir+"\\"+state_name+"\\*.xlsx"
    state_name=normalize_and_replace(state_name)
 

    for xls_path in glob.glob(state_path_xls)+glob.glob(state_path_xlsx):
        sho+=1
        kk+=1
        if sho==1:
            name_exist_in_sheet=find_state_name_in_sheet(xls_path,state_name)
            alram_string="[INFO] We did not find"+state_name+"!"
            good_alram_string="[INFO] We  found"+state_name+"!"
            if name_exist_in_sheet=="":
                
                print(f"\033[31m{alram_string}\033[0m")
            else:
                print(f"\033[34m{good_alram_string}\033[0m")
            if name_exist_in_sheet!="":
                
                # df = pd.read_excel(xls_path, engine='openpyxl',usecols=list(range(0,10)),sheet_name=name_exist_in_sheet,dtype=column_dtype)
                print(name_exist_in_sheet,"dadasdsa")
                print(xls_path)
                df = pd.read_excel(xls_path,sheet_name=name_exist_in_sheet,usecols=[8],skiprows=3,nrows=75,dtype=str)
                datas=df.fillna("").values.flatten()
                print(datas,name_exist_in_sheet)
                row_index=condition_check(df1,name_exist_in_sheet)
                if row_index!=None:
                    for i in range(75):
                        main_cell=ws.cell(row=row_index+2,column=3+i,value=to_number(datas[i]))
                        main_cell.fill=neon_green_fill
                else:
                    alram_string="[INFO] Finally!! We did not find"+state_name+"!"
                    print(f"\033[31m{alram_string}\033[0m")
        else:
            
            break
    if (sho==0):
        alram_string="[INFO] There is no file in folder"+state_name+"!"
        print(f"\033[31m{alram_string}\033[0m")

wb.save(output_path)

